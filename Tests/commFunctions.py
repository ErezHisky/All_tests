import openpyxl
from Drivers.DigAtt64bit import *
from Drivers.StepEngineCtrl import *
from Drivers.SigGenCtrl import *
from Drivers.ConnectToWifi import *
from Drivers.Terminal.SendMsg import *
from Drivers.Terminal.SendDataMsg import *
from Drivers.Terminal.iCDs import *
import math


def ipusAll():
    DigAtt_putMax()
    #signal_off()
    checkWifi()
    Engine_GoToHome()
    Hub_Gold_Unit_off()

def checkWifi():
    # check wifi to the unit
    myWifi = Wifi_Connection()
    try:
        print("trying to connect to wifi...")
        net = myWifi.wifi_connect_to_terminal()
        if net == None:
            print("couldn't connect to unit wifi, trying again...")
            checkWifi()
        else:
            print(f"connected to {net}")
    except:
        pass

def DigAtt_putMax():
    myAtt = DigAtt64()
    myAtt.set_attenuation(58.0)

def signal_off():
    try:
        mySig = MySignalGenerator()
        mySig.turnModOff()
        mySig.turnRFOff()
        mySig.closeConnection()
    except:
        print("turn signal generator on... and press yes")
        ans = input()
        if ans.lower() == "yes":
            signal_off()
        else:
            print("press yes when told...")

def Engine_GoToHome():
    my_eng = MyStepEngine("azimuth")
    my_eng.check_communication()
    my_eng.get_to_home()
    time.sleep(1)
    my_eng.engineType = "elevation"
    my_eng.get_to_home()
    my_eng.close_engine()

def Hub_Gold_Unit_off():
    pass

def getHRprops(unitNumber):
    wb = openpyxl.load_workbook("D:/ExcelFiles/RH_prop_example.xlsx")
    ws = wb.active
    for val in range(2, ws.max_row + 1):
        if str(ws['A' + str(val)].value) == unitNumber:
            return [ws['B'+str(val)].value, ws['C'+str(val)].value]

def CheckPingAllAngles(host, n=0):
    from multiping import MultiPing

    if (n > 0):
        avg = 0
        for i in range(n):
            avg += CheckPingAllAngles(host)
        avg = avg / n
    # Create a MultiPing object to test hosts / addresses
    mp = MultiPing([host])

    # Send the pings to those addresses
    mp.send()

    # With a 1 second timout, wait for responses (may return sooner if all
    # results are received).
    responses, no_responses = mp.receive(1)

    for addr, rtt in responses.items():
        RTT = rtt

    if no_responses:
        # Sending pings once more, but just to those addresses that have not
        # responded, yet.
        mp.send()
        responses, no_responses = mp.receive(1)
        RTT = -1

    return RTT

all_strcks = icd_strucks()
my_strFormats = all_strcks.strFormats
sdm = Send_The_Data()

def turn_tx_off():
    # turn Tx off
    print("turning tx off...")
    all_strcks.MxFEAxiRegMsg["cmd"] = all_strcks.iCD.SPI_WRITE
    all_strcks.MxFEAxiRegMsg["addr"] = 0x17
    all_strcks.MxFEAxiRegMsg["data"] = 0x3
    sdm.send_the_message("MxFEAxiRegMsg", all_strcks.MxFEAxiRegMsg)
    time.sleep(0.5)

def turn_tx_on():
    # turn Tx off
    print("turning tx on...")
    all_strcks.MxFEAxiRegMsg["cmd"] = all_strcks.iCD.SPI_WRITE
    all_strcks.MxFEAxiRegMsg["addr"] = 0x17
    all_strcks.MxFEAxiRegMsg["data"] = 0x7
    sdm.send_the_message("MxFEAxiRegMsg", all_strcks.MxFEAxiRegMsg)
    time.sleep(0.5)

def open_Tx_pic():
    all_strcks.PICReadWriteMsg["cmd"] = all_strcks.iCD.SPI_WRITE
    all_strcks.PICReadWriteMsg["addr"] = 1
    all_strcks.PICReadWriteMsg["data"] = 0
    sdm.send_the_message("PICReadWriteMsg", all_strcks.PICReadWriteMsg)
    time.sleep(0.5)

def gap_between_tx_cal(gap=1e6):
    val = (gap * math.pow(2,32)) / (40 * 1e6)
    all_strcks.modemRegCommandMsg["opcode"] = all_strcks.iCD.MODEM_REGISTER
    all_strcks.modemRegCommandMsg["addr"] = int(0x50)
    all_strcks.modemRegCommandMsg["data"] = int(val)
    sdm.send_the_message("modemRegCommandMsg", all_strcks.modemRegCommandMsg)
    time.sleep(0.5)

def set_TPC(tpc_indx):
    tpc71val = TPC_VEC[tpc_indx][0]
    tpcPIC4val = TPC_VEC[tpc_indx][1]
    #print(f"tpc71val is {tpc71val} tpcPIC4val is {tpcPIC4val}")
    # send the tpc message:
    all_strcks.PICReadWriteMsg["cmd"] = all_strcks.iCD.SPI_WRITE
    all_strcks.PICReadWriteMsg["addr"] = 4
    all_strcks.PICReadWriteMsg["data"] = tpcPIC4val
    sdm.send_the_message("PICReadWriteMsg", all_strcks.PICReadWriteMsg)
    time.sleep(0.2)
    all_strcks.MxFEAxiRegMsg["cmd"] = all_strcks.iCD.SPI_WRITE
    all_strcks.MxFEAxiRegMsg["addr"] = 0x71
    all_strcks.MxFEAxiRegMsg["data"] = tpc71val
    sdm.send_the_message("MxFEAxiRegMsg", all_strcks.MxFEAxiRegMsg)
    time.sleep(0.2)

def set_i_q(i_val=0x246f, q_val=0x246f):
    all_strcks.MxFEAxiRegMsg["cmd"] = all_strcks.iCD.SPI_WRITE
    all_strcks.MxFEAxiRegMsg["addr"] = 0x6f
    all_strcks.MxFEAxiRegMsg["data"] = i_val
    sdm.send_the_message("MxFEAxiRegMsg", all_strcks.MxFEAxiRegMsg)
    time.sleep(0.2)
    all_strcks.MxFEAxiRegMsg["cmd"] = all_strcks.iCD.SPI_WRITE
    all_strcks.MxFEAxiRegMsg["addr"] = 0x70
    all_strcks.MxFEAxiRegMsg["data"] = q_val
    sdm.send_the_message("MxFEAxiRegMsg", all_strcks.MxFEAxiRegMsg)
    time.sleep(0.2)

def set_cw():
    print("turning cw on...")
    all_strcks.MxFEAxiRegMsg["addr"] = 0x6e
    all_strcks.MxFEAxiRegMsg["data"] = 0x1
    sdm.send_the_message("MxFEAxiRegMsg", all_strcks.MxFEAxiRegMsg)
    time.sleep(0.2)

def read_i():
    pass

def read_q():
    pass

def calculate_p1db(real_line):
    p1db = [] # point, value
    p1dbp = 100
    p1 = p2 = 0
    delta = []
    ther = 0
    ther_line = [] # theoretical line

    for pout, a in real_line:
        p2 = p1
        p1 = pout
        delta.append(p2 - p1)

    for a in range(2, 77):
        if (abs(delta[a]) > 1) and (abs(delta[a + 1]) > 1):
            ther = a
            break

    for b in range(77):
        ther_line.append(real_line[ther][0] - (ther + 1 - b))

    min_d = []
    for c in range(77):
        min_d.append(abs(ther_line[c] - real_line[c][0]))

    # find min value place in the array:
    minmind = min(min_d)
    print(minmind)
    for d in range(len(min_d)):
        if min_d[d] == minmind:
            p1dbp = d
            break

    p1db.append(real_line[p1dbp][0]) #power
    p1db.append(real_line[p1dbp][1]) #tpc_num

    return p1db